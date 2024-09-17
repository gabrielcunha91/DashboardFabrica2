import streamlit as st
import pandas as pd
from babel.dates import format_date
from utils.queries import *
from utils.functions.cmv import *
from utils.functions.dados_gerais import *
from utils.functions.faturamento_zig import *
from utils.components import *
from utils.user import logout
import openpyxl
from dateutil.parser import parse
from pandas.tseries.offsets import MonthEnd, MonthBegin


st.set_page_config(
  layout = 'wide',
  page_title = 'DRE',
  page_icon=':⚖',
  initial_sidebar_state="collapsed"
)  
pd.set_option('future.no_silent_downcasting', True)

if 'loggedIn' not in st.session_state or not st.session_state['loggedIn']:
  st.switch_page('Login.py')

lojasComDados = preparar_dados_lojas_user()
data_inicio_default, data_fim_default = preparar_dados_datas()

col1, col2, col3 = st.columns([2, 1, 1])
# Adiciona seletores
with col1:
    lojas_selecionadas = st.multiselect(label='Selecione a loja dos dados a serem carregados', options=lojasComDados, key='lojas_multiselect')
with col2:
    data_inicio = st.date_input('Data de Início dos dados', value=data_inicio_default, key='data_inicio_input', format="DD/MM/YYYY")
with col3:
    data_fim = st.date_input('Data de Fim dos dados', value=data_fim_default, key='data_fim_input', format="DD/MM/YYYY")

# Converte as datas selecionadas para o formato Timestamp
data_inicio = pd.to_datetime(data_inicio)
data_fim = pd.to_datetime(data_fim)


# Dicionário para armazenar os resultados com nomes dinâmicos
resultados_orcamento = {}
totais_orcamento = {}  # Dicionário para armazenar as linhas "Total"

# Loop para cada mês entre data_inicio e data_fim
mes_atual = data_inicio
contador = 1  # Para criar nomes como OrcamentoFaturamento1, OrcamentoFaturamento2, etc.

while mes_atual <= data_fim:
    # Define o primeiro e último dia do mês atual
    inicio_mes = mes_atual.replace(day=1)
    fim_mes = mes_atual + MonthEnd(1)
    
    # Chama a função para o mês atual
    OrcamentoFaturamento = config_orcamento_faturamento(lojas_selecionadas, inicio_mes, fim_mes)
    
    # Salva o resultado no dicionário com um nome dinâmico
    nome_variavel = f'OrcamentoFaturamento{contador}'
    resultados_orcamento[nome_variavel] = OrcamentoFaturamento
    
    # Avança para o próximo mês
    mes_atual = mes_atual + MonthBegin(1)
    contador += 1

# Exemplo de como acessar e mostrar os resultados usando Streamlit
for nome, resultado in resultados_orcamento.items():
    # Manipula o DataFrame removendo colunas e ajustando valores
    resultado.drop(['Desconto', 'Valor Líquido'], axis=1, errors='ignore', inplace=True)
    resultado.loc[resultado['Orçamento'] == '0,00', 'Atingimento %'] = '-'
    
    # Seleciona a linha onde 'Categoria' é 'Total' e salva em um dicionário separado
    linha_total = resultado[resultado['Categoria'] == 'Total Geral']
    totais_orcamento[f'{nome}_Total'] = linha_total

    resultado.drop(['Atingimento %', 'Faturam - Orçamento'], axis=1, errors='ignore', inplace=True)

    st.write(f"{nome}:")
    st.write(resultado)

# Exibe as linhas "Total" separadamente
st.write("Valores Totais por Mês:")
for nome_total, total in totais_orcamento.items():
    st.write(f"{nome_total}:")
    st.write(total)


def extrair_descontos_por_mes(caminho_arquivo, nome_planilha):
    # Carrega o arquivo e a planilha específica
    workbook = openpyxl.load_workbook(caminho_arquivo, data_only=True)
    sheet = workbook[nome_planilha]

    valores_total = []

    for row in sheet.iter_rows(min_row=1):
        cell = row[1]
        # Procura pela palavra "Total" ou "TOTAL" na segunda coluna, apenas se o valor for uma string
        if isinstance(cell.value, str) and cell.value.strip().lower() == "desconto final":
            valor_total = sheet.cell(row=cell.row, column=3).value  # Pega o valor ao lado (coluna C)
            valor_total = format_brazilian(valor_total)
            valores_total.append({'Desconto_por_mes': valor_total})

    return pd.DataFrame(valores_total)


aliquotas_tributacao = {
    'Alíquota de Tributação - SIMPLES': 0.03,
    'Alíquota de Tributação - SIMPLES (cont. param.)': 0.07,
    'PIS/COFINS': 0.0365,
    'IRPJ (presumido)': 0.012,
    'CSLL (presumido)': 0.0108,
    'CPMF': 0.0,
    'ICMS': 0.032,
    'ISS': 0.05
}



def processar_total_acesso(caminho_arquivo):
    # Carrega o arquivo Excel
    workbook = openpyxl.load_workbook(caminho_arquivo, data_only=True)
    
    # Verifica se a aba "Total Acesso" existe
    if "Total Acesso" in workbook.sheetnames:
        # Carrega a aba "Total Acesso" em um DataFrame
        df_total_acesso = pd.read_excel(caminho_arquivo, sheet_name="Total Acesso")
        
        # Verifica se as colunas necessárias existem no DataFrame
        if {'ProductDescription', 'Value', 'Competência'}.issubset(df_total_acesso.columns):
            # Agrupa por 'ProductDescription' e 'Competência', somando os valores de 'Value'
            df_agrupado = df_total_acesso.groupby(['ProductDescription', 'Competência'], as_index=False)['Value'].sum()
            return df_agrupado
        else:
            return "Colunas 'ProductDescription', 'Value' ou 'Competência' não encontradas na aba 'Total Acesso'."
    else:
        return "A aba 'Total Acesso' não foi encontrada."
    


# def processar_total_acesso(caminho_arquivo):
#     # Carrega o arquivo Excel
#     workbook = openpyxl.load_workbook(caminho_arquivo, data_only=True)
    
#     # Verifica se a aba "Total Acesso" existe
#     if "Total Acesso" in workbook.sheetnames:
#         # Carrega a aba "Total Acesso" em um DataFrame
#         df_total_acesso = pd.read_excel(caminho_arquivo, sheet_name="Total Acesso")
        
#         # Verifica se as colunas necessárias existem no DataFrame
#         if {'ProductDescription', 'Value', 'Competência'}.issubset(df_total_acesso.columns):
#             # Ajusta a coluna 'Competência' para manter o formato mes/ano
#             df_total_acesso['Competência'] = pd.to_datetime(df_total_acesso['Competência'], errors='coerce').dt.strftime('%b/%y')
            
#             # Agrupa por 'ProductDescription' e 'Competência', somando os valores de 'Value'
#             df_agrupado = df_total_acesso.groupby(['ProductDescription', 'Competência'], as_index=False)['Value'].sum()
#             return df_agrupado
#         else:
#             return "Colunas 'ProductDescription', 'Value' ou 'Competência' não encontradas na aba 'Total Acesso'."
#     else:
#         return "A aba 'Total Acesso' não foi encontrada."



# Exemplo de uso no Streamlit
uploaded_file = st.file_uploader("Faça o upload do arquivo Excel", type=["xlsx"])

if uploaded_file is not None:
    df_valores_total = extrair_descontos_por_mes(uploaded_file, 'Descontos')
    st.write("Descontos por Mês:")
    st.dataframe(df_valores_total)

    resultado_total_acesso = processar_total_acesso(uploaded_file)
    
    # Verifica se o resultado é um DataFrame ou uma mensagem de erro
    if isinstance(resultado_total_acesso, pd.DataFrame):
        st.write("Dados agrupados da aba 'Total Acesso':")
        st.dataframe(resultado_total_acesso)
    else:
        st.write(resultado_total_acesso)