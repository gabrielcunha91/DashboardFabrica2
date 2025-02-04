import streamlit as st
import pandas as pd
import numpy as np
from utils.queries import *
from workalendar.america import Brazil
from utils.functions.dados_gerais import *
import openpyxl
import os

def convert_to_datetime(df, cols):
  for col in cols:
    df[col] = pd.to_datetime(df[col])
  return df


def filtrar_data_fim(dataframe, data_fim, categoria):
  data_fim = pd.Timestamp(data_fim)
  dataframe.loc[:, categoria] = pd.to_datetime(dataframe[categoria], errors='coerce')
  dataframe_filtered = dataframe.loc[
    (dataframe[categoria] <= data_fim)
  ]
  return dataframe_filtered


def prolongar_projecao(df_projecao_zig, dias_prolongados=7):
  df_projecao_zig = df_projecao_zig.copy()
  df_projecao_zig['Data'] = pd.to_datetime(df_projecao_zig['Data'])
  novas_linhas = []
  
  for empresa, grupo in df_projecao_zig.groupby('Empresa'):
    for i in range(1, dias_prolongados + 1):
      df_temp = grupo.copy()
      df_temp['Data'] += pd.Timedelta(days=7)  # Avança 7 dias mantendo padrão de semana
      novas_linhas.append(df_temp)

  df_projecao_estendida = pd.concat([df_projecao_zig] + novas_linhas, ignore_index=True)
  df_projecao_estendida = df_projecao_estendida.sort_values(by=['Empresa', 'Data']).reset_index(drop=True)
  df_projecao_estendida.drop_duplicates(subset=['Data', 'Empresa'], keep='first', inplace=True)

  return df_projecao_estendida



def config_projecao_bares(multiplicador, data_fim):
  df_saldos_bancarios = GET_SALDOS_BANCARIOS()
  df_valor_liquido = GET_VALOR_LIQUIDO_RECEBIDO()
  df_projecao_zig = GET_PROJECAO_ZIG()
  df_receitas_extraord_proj = GET_RECEITAS_EXTRAORD_FLUXO_CAIXA()
  df_despesas_aprovadas = GET_DESPESAS_APROVADAS()
  df_despesas_pagas = GET_DESPESAS_PAGAS()

  df_projecao_zig = prolongar_projecao(df_projecao_zig, dias_prolongados=7)
  # Converter colunas de data
  dfs = [df_saldos_bancarios, df_valor_liquido, df_projecao_zig, df_receitas_extraord_proj, df_despesas_aprovadas, df_despesas_pagas]
  for df in dfs:
    df = convert_to_datetime(df, ['Data'])

  # Merge dataframes
  merged_df = df_saldos_bancarios
  for df in [df_valor_liquido, df_projecao_zig, df_receitas_extraord_proj, df_despesas_aprovadas, df_despesas_pagas]:
    merged_df = pd.merge(merged_df, df, on=['Data', 'Empresa'], how='outer')

  # Preenchendo valores nulos com 0 e renomeando colunas
  merged_df = merged_df.fillna(0)
  merged_df = merged_df.rename(columns={'Valor_Projetado': 'Valor_Projetado_Zig'})
  merged_df = merged_df.sort_values(by='Data').reset_index(drop=True)

  # Filtrando por data de fim
  merged_df = filtrar_data_fim(merged_df, data_fim, 'Data')

  # Ajustando formatação
  cols = ['Saldo_Inicio_Dia', 'Valor_Liquido_Recebido', 'Valor_Projetado_Zig', 'Receita_Projetada_Extraord', 'Despesas_Aprovadas_Pendentes', 'Despesas_Pagas']
  merged_df[cols] = merged_df[cols].astype(float).round(2)

  # Aplicando lógica de negócios
  merged_df['Valor_Projetado_Zig'] = merged_df.apply(lambda row: 0 if row['Valor_Liquido_Recebido'] > 0 else row['Valor_Projetado_Zig'], axis=1)
  merged_df['Valor_Projetado_Zig'] = merged_df['Valor_Projetado_Zig'] * multiplicador

  merged_df['Saldo_Final'] = merged_df['Saldo_Inicio_Dia'] + merged_df['Valor_Liquido_Recebido'] + merged_df['Valor_Projetado_Zig'] + merged_df['Receita_Projetada_Extraord'] - merged_df['Despesas_Aprovadas_Pendentes'] - merged_df['Despesas_Pagas']
  merged_df = format_date_brazilian(merged_df, 'Data')
  return merged_df

def is_in_group(empresa, houses_to_group):
  return any(house in empresa for house in houses_to_group)

def config_grouped_projecao(df_projecao_bares):
  houses_to_group = [
    'Bar Brahma - Centro', 'Bar Léo - Centro', 'Bar Brasilia -  Aeroporto ', 'Bar Brasilia -  Aeroporto', 'Delivery Bar Leo Centro', 
    'Delivery Fabrica de Bares', 'Delivery Orfeu', 'Edificio Rolim', 'Hotel Maraba', 
    'Jacaré', 'Orfeu', 'Riviera Bar', 'Tempus', 'Escritório Fabrica de Bares', 'Priceless', 'Girondino - CCBB', 'Girondino ', 'Bar Brahma - Granja', 'Edificio Rolim'
  ]

  grouped_df = df_projecao_bares[df_projecao_bares['Empresa'].apply(lambda x: is_in_group(x, houses_to_group))]
  grouped_df = grouped_df.groupby(['Data']).sum().reset_index()
  return grouped_df


def config_feriados():
  calendario_brasil = Brazil()
  anos_desejados = list(range(2023, 2031))
  datas_feriados = []

  for ano in anos_desejados:
    feriados_ano = calendario_brasil.holidays(ano)
    datas_feriados.extend([feriado[0] for feriado in feriados_ano])

  serie_datas_feriados = pd.Series(datas_feriados, name='Data_Feriado')
  serie_datas_feriados = pd.to_datetime(serie_datas_feriados)
  nova_data = pd.to_datetime('2024-03-29')
  serie_datas_feriados = pd.concat([serie_datas_feriados, pd.Series([nova_data])]).reset_index(drop=True)

  return serie_datas_feriados


def calcular_taxa(row, taxas):
  if row['Tipo_Pagamento'] == 'DÉBITO':
    return row['Valor_Faturado'] * taxas['DÉBITO']
  elif row['Tipo_Pagamento'] == 'CRÉDITO' and row['Antecipacao_Credito'] == 1:
    return row['Valor_Faturado'] * taxas['CRÉDITO_ANTECIPADO']
  elif row['Tipo_Pagamento'] == 'CRÉDITO' and row['Antecipacao_Credito'] == 0:
    return row['Valor_Faturado'] * taxas['CRÉDITO_PADRAO']
  elif row['Tipo_Pagamento'] == 'APP':
    return row['Valor_Faturado'] * taxas['APP']
  elif row['Tipo_Pagamento'] == 'PIX':
    return row['Valor_Faturado'] * taxas['PIX']
  else:
    return 0.00

def ajustar_data_compensacao(row, serie_datas_feriados):
  if row['Tipo_Pagamento'] == 'DÉBITO':
    row['Data_Compensacao'] += pd.Timedelta(days=1)
  elif row['Tipo_Pagamento'] == 'CRÉDITO' and row['Antecipacao_Credito'] == 1:
    row['Data_Compensacao'] += pd.Timedelta(days=1)
#########################     AQUI ESTÁ O CRÉDITO     ###############################
  elif row['Tipo_Pagamento'] == 'CRÉDITO' and row['Antecipacao_Credito'] == 0:
    row['Data_Compensacao'] += pd.Timedelta(days=31)
    ################################################
  elif row['Tipo_Pagamento'] in ['PIX', 'DINHEIRO', 'VOUCHER', 'ANTECIPADO']:
    row['Data_Compensacao'] += pd.Timedelta(days=1)
    #################### APP CONTA IGUAL CRÉDITO ########################
  elif row['Tipo_Pagamento'] == 'APP':
    row['Data_Compensacao'] += pd.Timedelta(days=31)

  if row['Data_Compensacao'] in serie_datas_feriados.values:
    row['Data_Compensacao'] += pd.Timedelta(days=1)

  if row['Data_Compensacao'].strftime('%A') == 'Sunday':
    row['Data_Compensacao'] += pd.Timedelta(days=1)
  elif row['Data_Compensacao'].strftime('%A') == 'Saturday':
    row['Data_Compensacao'] += pd.Timedelta(days=2)

  if row['Data_Compensacao'] in serie_datas_feriados.values:
    row['Data_Compensacao'] += pd.Timedelta(days=1)

  return row['Data_Compensacao']

def config_faturam_zig_fluxo_caixa(serie_datas_feriados):
  taxas = {
    'DÉBITO': 0.0095,
    'CRÉDITO_ANTECIPADO': 0.0265,
    'CRÉDITO_PADRAO': 0.016,
    'APP': 0.0074,
    'PIX': 0.0074,
  }

  lojas_sem_antecipacao = ['Arcos', 'Bar Léo - Centro', 'Blue Note - São Paulo', 'Blue Note SP (Novo)', 'Jacaré', 'Love Cabaret', 'Orfeu']

  df_faturam_zig = GET_FATURAMENTO_ZIG_FLUXO_CAIXA()
  df_faturam_zig['Data_Faturamento'] = pd.to_datetime(df_faturam_zig['Data_Faturamento'])
  df_faturam_zig['Valor_Faturado'] = df_faturam_zig['Valor_Faturado'].astype(float).round(2)

  df_faturam_zig['Antecipacao_Credito'] = df_faturam_zig.apply(lambda row: 0 if row['Loja'] in lojas_sem_antecipacao else 1, axis=1)
  df_faturam_zig['Taxa'] = df_faturam_zig.apply(lambda row: calcular_taxa(row, taxas), axis=1)

  df_faturam_zig['Valor_Compensado'] = df_faturam_zig.apply(
    lambda row: 0 if row['Tipo_Pagamento'] == 'BÔNUS' else row['Valor_Faturado'] - row['Taxa'], axis=1)

  df_faturam_zig['Custos_Zig'] = 0.008 * df_faturam_zig['Valor_Faturado']
  df_faturam_zig['Accumulated_Cost'] = df_faturam_zig.groupby(
    [df_faturam_zig['Data_Faturamento'].dt.year, df_faturam_zig['Data_Faturamento'].dt.month, df_faturam_zig['ID_Loja']]
  )['Custos_Zig'].cumsum()

  exceeded_limit = df_faturam_zig['Accumulated_Cost'] >= 2800.00
  df_faturam_zig.loc[exceeded_limit, 'Custos_Zig'] = np.maximum(0, 2800.00 - (df_faturam_zig['Accumulated_Cost'] - df_faturam_zig['Custos_Zig']))
  df_faturam_zig.loc[exceeded_limit, 'Accumulated_Cost'] = np.minimum(2800.00, df_faturam_zig.loc[exceeded_limit, 'Accumulated_Cost'])

  df_faturam_zig = df_faturam_zig.drop('Accumulated_Cost', axis=1)
  df_faturam_zig['Valor_Final'] = df_faturam_zig.apply(
    lambda row: row['Custos_Zig'] * (-1) if row['Tipo_Pagamento'] in ['VOUCHER', 'DINHEIRO'] else row['Valor_Compensado'] - row['Custos_Zig'],
    axis=1
  )

  df_faturam_zig = df_faturam_zig.round({'Taxa': 2, 'Valor_Compensado': 2, 'Custos_Zig': 2, 'Valor_Final': 2})

  df_faturam_zig['Data_Compensacao'] = df_faturam_zig['Data_Faturamento']
  df_faturam_zig['Data_Compensacao'] = df_faturam_zig.apply(lambda row: ajustar_data_compensacao(row, serie_datas_feriados), axis=1)

  df_faturam_zig['Data_Compensacao'] = df_faturam_zig['Data_Compensacao'].dt.date
  df_faturam_zig['Data_Faturamento'] = df_faturam_zig['Data_Faturamento'].dt.date

  return df_faturam_zig





def config_receitas_extraord_fluxo_caixa():
  df_receitas_extraord = GET_RECEITAS_EXTRAORD_CONCILIACAO()
  cols = ['Data_Competencia', 'Data_Venc_Parc_1', 'Data_Receb_Parc_1', 'Data_Venc_Parc_2', 'Data_Receb_Parc_2', 'Data_Venc_Parc_3',
          'Data_Receb_Parc_3', 'Data_Venc_Parc_4', 'Data_Receb_Parc_4', 'Data_Venc_Parc_5', 'Data_Receb_Parc_5']
  
  for col in cols:
    df_receitas_extraord[col] = pd.to_datetime(df_receitas_extraord[col]).dt.date

  cols = ['Valor_Total', 'Categ_AB', 'Categ_Aluguel', 'Categ_Artist', 'Categ_Couvert', 'Categ_Locacao',
          'Categ_Patroc', 'Categ_Taxa_Serv', 'Valor_Parc_1', 'Valor_Parc_2', 'Valor_Parc_3', 'Valor_Parc_4', 'Valor_Parc_5']
  
  for col in cols:
    df_receitas_extraord[col] = df_receitas_extraord[col].astype(float).round(2)

  return df_receitas_extraord


def config_view_parc_agrup():
  df_view_parc_agrup = GET_VIEW_PARC_AGRUP()
  df_view_parc_agrup = df_view_parc_agrup.drop('Numero_Linha', axis=1)
  df_view_parc_agrup['Data_Vencimento'] = df_view_parc_agrup['Data_Vencimento'].dt.date
  df_view_parc_agrup['Data_Recebimento'] = df_view_parc_agrup['Data_Recebimento'].dt.date
  df_view_parc_agrup['Data_Ocorrencia'] = df_view_parc_agrup['Data_Ocorrencia'].dt.date
  df_view_parc_agrup['Valor_Parcela'] = df_view_parc_agrup['Valor_Parcela'].astype(float).round(2)
  return df_view_parc_agrup



def config_custos_blueme_sem_parcelamento():
  df_custos_blueme_sem_parcelamento = GET_CUSTOS_BLUEME_SEM_PARCELAMENTO()
  df_custos_blueme_sem_parcelamento['Valor'] = df_custos_blueme_sem_parcelamento['Valor'].astype(float).round(2)
  
  date_columns = ['Data_Vencimento', 'Data_Competencia', 'Data_Lancamento', 'Realizacao_Pgto', 'Previsao_Pgto']
  
  for col in date_columns:
    df_custos_blueme_sem_parcelamento[col] = pd.to_datetime(df_custos_blueme_sem_parcelamento[col], errors='coerce')
  
  return df_custos_blueme_sem_parcelamento
 


def config_custos_blueme_com_parcelamento():
  df_custos_blueme_com_parcelamento = GET_CUSTOS_BLUEME_COM_PARCELAMENTO()
  df_custos_blueme_com_parcelamento['Valor_Parcela'] = df_custos_blueme_com_parcelamento['Valor_Parcela'].astype(float).round(2)
  df_custos_blueme_com_parcelamento['Valor_Original'] = df_custos_blueme_com_parcelamento['Valor_Original'].astype(float).round(2)
  df_custos_blueme_com_parcelamento['Valor_Liquido'] = df_custos_blueme_com_parcelamento['Valor_Liquido'].astype(float).round(2)
  
  date_columns = ['Vencimento_Parcela', 'Previsao_Parcela', 'Realiz_Parcela', 'Data_Lancamento']
  
  for col in date_columns:
    df_custos_blueme_com_parcelamento[col] = pd.to_datetime(df_custos_blueme_com_parcelamento[col], format='%d/%m/%Y', errors='coerce')
  
  return df_custos_blueme_com_parcelamento


def config_extratos():
  df_extratos = GET_EXTRATOS_BANCARIOS()
  df_extratos['Data_Transacao'] = df_extratos['Data_Transacao'].dt.date
  df_extratos['Valor'] = df_extratos['Valor'].astype(float).round(2)
  return df_extratos


def config_mutuos():
  df_mutuos = GET_MUTUOS()
  df_mutuos['Data_Mutuo'] = df_mutuos['Data_Mutuo'].dt.date
  df_mutuos['Valor'] = df_mutuos['Valor'].astype(float).round(2)
  return df_mutuos

def config_tesouraria_trans():
  df_tesouraria_trans = GET_TESOURARIA_TRANSACOES()
  df_tesouraria_trans['Data_Transacao'] = df_tesouraria_trans['Data_Transacao'].dt.date
  df_tesouraria_trans['Valor'] = df_tesouraria_trans['Valor'].astype(float).round(2)
  return df_tesouraria_trans

def config_ajustes_conciliacao():
  df_ajustes_conciliacao = GET_AJUSTES_CONCILIACAO()
  df_ajustes_conciliacao['Data Ajuste'] = df_ajustes_conciliacao['Data Ajuste'].dt.date
  df_ajustes_conciliacao['Valor'] = df_ajustes_conciliacao['Valor'].astype(float).round(2)
  return df_ajustes_conciliacao

def somar_total(df):
  colunas_numericas = df.select_dtypes(include=[int, float]).columns
  soma_colunas = df[colunas_numericas].sum().to_frame().T
  soma_colunas['Data'] = 'Total' 
  df_com_soma = pd.concat([df, soma_colunas], ignore_index=True)
  return df_com_soma



def config_despesas_a_pagar(lojas_selecionadas, dataInicio, dataFim):
  despesasPendentes = GET_DESPESAS_PENDENTES(dataInicio=dataInicio, dataFim=dataFim)
  despesasPendentes = filtrar_por_classe_selecionada(despesasPendentes, 'Loja', lojas_selecionadas)
  return despesasPendentes



def export_to_excel(df, sheet_name, excel_filename):
  if os.path.exists(excel_filename):
    wb = openpyxl.load_workbook(excel_filename)
  else:
    wb = openpyxl.Workbook()

  if sheet_name in wb.sheetnames:
    wb.remove(wb[sheet_name])
  ws = wb.create_sheet(title=sheet_name)
  
  # Escrever os cabeçalhos
  for col_idx, column_title in enumerate(df.columns, start=1):
    ws.cell(row=1, column=col_idx, value=column_title)
  
  # Escrever os dados
  for row_idx, row in enumerate(df.itertuples(index=False, name=None), start=2):
    for col_idx, value in enumerate(row, start=1):
      ws.cell(row=row_idx, column=col_idx, value=value)

  wb.save(excel_filename)